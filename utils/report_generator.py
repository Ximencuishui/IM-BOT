"""
报表生成工具
- Excel报表生成
- 按线路/销售分组汇总
"""
import os
import logging
from datetime import date, datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config.settings import settings

logger = logging.getLogger(__name__)


class ReportGenerator:
    """报表生成器"""

    def __init__(self):
        self.output_dir = 'reports'
        os.makedirs(self.output_dir, exist_ok=True)

        # 样式定义
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        self.title_font = Font(bold=True, size=16)
        self.title_alignment = Alignment(horizontal="center", vertical="center")

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_daily_report(self, orders: List[Dict], report_date: date) -> str:
        """
        生成日报表

        Args:
            orders: 订单列表
            report_date: 报表日期

        Returns:
            Excel文件路径
        """
        filename = f"配送报表_{report_date.isoformat()}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()

        # Sheet 1: 订单明细
        ws_detail = wb.active
        ws_detail.title = "订单明细"
        self._write_detail_sheet(ws_detail, orders, report_date)

        # Sheet 2: 按线路汇总
        ws_route = wb.create_sheet("按线路汇总")
        self._write_route_summary(ws_route, orders, report_date)

        # Sheet 3: 按销售汇总
        ws_sales = wb.create_sheet("按销售汇总")
        self._write_sales_summary(ws_sales, orders, report_date)

        # Sheet 4: 商品汇总
        ws_product = wb.create_sheet("商品汇总")
        self._write_product_summary(ws_product, orders, report_date)

        wb.save(filepath)
        logger.info(f"报表生成成功: {filepath}")

        return filepath

    def _write_detail_sheet(self, ws, orders: List[Dict], report_date: date):
        """写入订单明细表"""
        # 标题
        ws.merge_cells('A1:J1')
        title_cell = ws.cell(row=1, column=1, value=f"配送报表 - {report_date.isoformat()}")
        title_cell.font = self.title_font
        title_cell.alignment = self.title_alignment

        # 表头
        headers = ['序号', '客户名称', '联系电话', '配送地址', '商品名称',
                   '数量', '单位', '单价', '小计', '备注']
        ws.append(headers)

        # 应用表头样式
        for cell in ws[2]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 数据行
        row_num = 3
        for idx, order in enumerate(orders, 1):
            customer = order.get('customer', {}) or {}
            customer_name = order.get('customer_name', '')
            phone = customer.get('phone', '')
            address = customer.get('address', '')
            remark = order.get('remark', '')

            for item_idx, item in enumerate(order.get('items', [])):
                ws.cell(row=row_num, column=1, value=idx if item_idx == 0 else '')
                ws.cell(row=row_num, column=2, value=customer_name if item_idx == 0 else '')
                ws.cell(row=row_num, column=3, value=phone if item_idx == 0 else '')
                ws.cell(row=row_num, column=4, value=address if item_idx == 0 else '')
                ws.cell(row=row_num, column=5, value=item.get('product_name', ''))
                ws.cell(row=row_num, column=6, value=float(item.get('quantity', 0)))
                ws.cell(row=row_num, column=7, value=item.get('unit', '斤'))
                ws.cell(row=row_num, column=8, value=float(item.get('unit_price', 0)))
                ws.cell(row=row_num, column=9, value=float(item.get('subtotal', 0)))
                item_remark = item.get('remark', '')
                ws.cell(row=row_num, column=10, value=item_remark or (remark if item_idx == 0 else ''))

                # 应用边框
                for col in range(1, 11):
                    ws.cell(row=row_num, column=col).border = self.border

                row_num += 1

        # 合计行
        total_amount = sum(o['total_amount'] for o in orders)
        ws.merge_cells(f'A{row_num}:I{row_num}')
        total_cell = ws.cell(row=row_num, column=1, value='合计')
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=10, value=round(total_amount, 2)).font = Font(bold=True)

        # 调整列宽
        column_widths = [8, 15, 15, 25, 12, 10, 8, 10, 12, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _write_route_summary(self, ws, orders: List[Dict], report_date: date):
        """写入按线路汇总表"""
        # 标题
        ws.merge_cells('A1:E1')
        title_cell = ws.cell(row=1, column=1, value=f"按线路汇总 - {report_date.isoformat()}")
        title_cell.font = self.title_font
        title_cell.alignment = self.title_alignment

        # 表头
        headers = ['线路名称', '订单数', '客户数', '总金额', '占比']
        ws.append(headers)

        for cell in ws[2]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 统计数据
        route_stats = {}
        total_amount = 0

        for order in orders:
            customer = order.get('customer', {}) or {}
            route = customer.get('route', {}) or {}
            route_name = route.get('route_name', '未分配')

            if route_name not in route_stats:
                route_stats[route_name] = {
                    'order_count': 0,
                    'customers': set(),
                    'total_amount': 0.0
                }

            route_stats[route_name]['order_count'] += 1
            route_stats[route_name]['customers'].add(order.get('customer_id'))
            route_stats[route_name]['total_amount'] += order['total_amount']
            total_amount += order['total_amount']

        # 写入数据
        row_num = 3
        for route_name, stats in sorted(route_stats.items()):
            percentage = (stats['total_amount'] / total_amount * 100) if total_amount > 0 else 0

            ws.cell(row=row_num, column=1, value=route_name)
            ws.cell(row=row_num, column=2, value=stats['order_count'])
            ws.cell(row=row_num, column=3, value=len(stats['customers']))
            ws.cell(row=row_num, column=4, value=round(stats['total_amount'], 2))
            ws.cell(row=row_num, column=5, value=f"{percentage:.1f}%")

            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = self.border

            row_num += 1

        # 合计
        ws.cell(row=row_num, column=1, value='合计').font = Font(bold=True)
        ws.cell(row=row_num, column=2, value=len(orders)).font = Font(bold=True)
        ws.cell(row=row_num, column=4, value=round(total_amount, 2)).font = Font(bold=True)
        ws.cell(row=row_num, column=5, value='100%').font = Font(bold=True)

        # 调整列宽
        for i, width in enumerate([20, 12, 12, 15, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _write_sales_summary(self, ws, orders: List[Dict], report_date: date):
        """写入按销售汇总表"""
        # 标题
        ws.merge_cells('A1:D1')
        title_cell = ws.cell(row=1, column=1, value=f"按销售汇总 - {report_date.isoformat()}")
        title_cell.font = self.title_font
        title_cell.alignment = self.title_alignment

        # 表头
        headers = ['销售人员', '订单数', '客户数', '总金额']
        ws.append(headers)

        for cell in ws[2]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 统计数据
        sales_stats = {}

        for order in orders:
            customer = order.get('customer', {}) or {}
            sales_person = customer.get('sales_person', '未分配')

            if sales_person not in sales_stats:
                sales_stats[sales_person] = {
                    'order_count': 0,
                    'customers': set(),
                    'total_amount': 0.0
                }

            sales_stats[sales_person]['order_count'] += 1
            sales_stats[sales_person]['customers'].add(order.get('customer_id'))
            sales_stats[sales_person]['total_amount'] += order['total_amount']

        # 写入数据
        row_num = 3
        for sales_person, stats in sorted(sales_stats.items()):
            ws.cell(row=row_num, column=1, value=sales_person)
            ws.cell(row=row_num, column=2, value=stats['order_count'])
            ws.cell(row=row_num, column=3, value=len(stats['customers']))
            ws.cell(row=row_num, column=4, value=round(stats['total_amount'], 2))

            for col in range(1, 5):
                ws.cell(row=row_num, column=col).border = self.border

            row_num += 1

        # 调整列宽
        for i, width in enumerate([15, 12, 12, 15], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _write_product_summary(self, ws, orders: List[Dict], report_date: date):
        """写入商品汇总表"""
        # 标题
        ws.merge_cells('A1:F1')
        title_cell = ws.cell(row=1, column=1, value=f"商品汇总 - {report_date.isoformat()}")
        title_cell.font = self.title_font
        title_cell.alignment = self.title_alignment

        # 表头
        headers = ['商品名称', '总数量', '单位', '平均单价', '总金额', '订单数']
        ws.append(headers)

        for cell in ws[2]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 统计数据
        product_stats = {}

        for order in orders:
            for item in order.get('items', []):
                product_name = item.get('product_name', '')

                if product_name not in product_stats:
                    product_stats[product_name] = {
                        'total_quantity': 0,
                        'unit': item.get('unit', '斤'),
                        'total_amount': 0.0,
                        'order_count': 0
                    }

                product_stats[product_name]['total_quantity'] += float(item.get('quantity', 0))
                product_stats[product_name]['total_amount'] += float(item.get('subtotal', 0))
                product_stats[product_name]['order_count'] += 1

        # 写入数据
        row_num = 3
        for product_name, stats in sorted(product_stats.items()):
            avg_price = stats['total_amount'] / stats['total_quantity'] if stats['total_quantity'] > 0 else 0

            ws.cell(row=row_num, column=1, value=product_name)
            ws.cell(row=row_num, column=2, value=round(stats['total_quantity'], 2))
            ws.cell(row=row_num, column=3, value=stats['unit'])
            ws.cell(row=row_num, column=4, value=round(avg_price, 2))
            ws.cell(row=row_num, column=5, value=round(stats['total_amount'], 2))
            ws.cell(row=row_num, column=6, value=stats['order_count'])

            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = self.border

            row_num += 1

        # 调整列宽
        for i, width in enumerate([15, 12, 10, 12, 12, 10], 1):
            ws.column_dimensions[get_column_letter(i)].width = width
