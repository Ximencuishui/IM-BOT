"""
报表管理API路由
- 日报表查询
- Excel导出
"""
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, date
import logging
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from services.order_service import order_service
from services.customer_service import customer_service

logger = logging.getLogger(__name__)

reports_bp = Blueprint('reports', __name__, url_prefix='/api/reports')


@reports_bp.route('/daily', methods=['GET'])
def get_daily_report():
    """
    获取日报表

    Query Params:
    - date: 日期 (YYYY-MM-DD), 默认今天
    - route_id: 线路ID(可选)
    """
    try:
        date_str = request.args.get('date')
        route_id = request.args.get('route_id', type=int)

        if date_str:
            report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            report_date = date.today()

        # 获取当日所有订单
        orders = order_service.get_orders_by_date_range(report_date, report_date)

        # 按线路/销售分组统计
        report_data = _generate_report_summary(orders, route_id)

        return jsonify({
            'success': True,
            'date': report_date.isoformat(),
            'summary': report_data
        }), 200

    except ValueError:
        return jsonify({'error': '日期格式错误,应为YYYY-MM-DD'}), 400
    except Exception as e:
        logger.error(f"查询日报表API异常: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/export', methods=['GET'])
def export_excel():
    """
    导出Excel报表

    Query Params:
    - date: 日期 (YYYY-MM-DD), 默认今天
    - route_id: 线路ID(可选)
    """
    try:
        date_str = request.args.get('date')
        route_id = request.args.get('route_id', type=int)

        if date_str:
            report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            report_date = date.today()

        # 获取订单数据
        orders = order_service.get_orders_by_date_range(report_date, report_date)

        # 生成Excel
        excel_file = _generate_excel(orders, report_date)

        # 返回文件
        filename = f"配送报表_{report_date.isoformat()}.xlsx"
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except ValueError:
        return jsonify({'error': '日期格式错误,应为YYYY-MM-DD'}), 400
    except Exception as e:
        logger.error(f"导出Excel API异常: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


def _generate_report_summary(orders: list, route_id: int = None) -> dict:
    """
    生成报表汇总数据

    Returns:
        {
            "total_orders": 10,
            "total_amount": 1234.56,
            "by_route": {...},
            "by_sales": {...},
            "order_details": [...]
        }
    """
    total_orders = len(orders)
    total_amount = sum(o['total_amount'] for o in orders)

    # 按线路分组
    by_route = {}
    # 按销售分组
    by_sales = {}

    for order in orders:
        customer = order.get('customer', {})
        route_name = customer.get('route', {}).get('route_name', '未分配')
        sales_person = customer.get('sales_person', '未分配')

        # 线路统计
        if route_name not in by_route:
            by_route[route_name] = {
                'order_count': 0,
                'total_amount': 0.0,
                'customers': []
            }
        by_route[route_name]['order_count'] += 1
        by_route[route_name]['total_amount'] += order['total_amount']
        by_route[route_name]['customers'].append(order['customer_name'])

        # 销售统计
        if sales_person not in by_sales:
            by_sales[sales_person] = {
                'order_count': 0,
                'total_amount': 0.0
            }
        by_sales[sales_person]['order_count'] += 1
        by_sales[sales_person]['total_amount'] += order['total_amount']

    return {
        'total_orders': total_orders,
        'total_amount': round(total_amount, 2),
        'by_route': by_route,
        'by_sales': by_sales,
        'order_details': orders
    }


def _generate_excel(orders: list, report_date: date) -> io.BytesIO:
    """生成Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"配送报表 {report_date.isoformat()}"

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = ['序号', '客户名称', '联系电话', '配送地址', '商品明细', '数量', '单位', '单价', '小计', '备注']
    ws.append(headers)

    # 应用表头样式
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 填充数据
    row_num = 2
    for idx, order in enumerate(orders, 1):
        customer_name = order.get('customer_name', '')
        remark = order.get('remark', '')

        # 合并单元格显示订单信息
        for item_idx, item in enumerate(order.get('items', [])):
            ws.cell(row=row_num, column=1, value=idx if item_idx == 0 else '')
            ws.cell(row=row_num, column=2, value=customer_name if item_idx == 0 else '')
            ws.cell(row=row_num, column=3, value=order.get('customer', {}).get('phone', '') if item_idx == 0 else '')
            ws.cell(row=row_num, column=4, value=order.get('customer', {}).get('address', '') if item_idx == 0 else '')
            ws.cell(row=row_num, column=5, value=item.get('product_name', ''))
            ws.cell(row=row_num, column=6, value=item.get('quantity', 0))
            ws.cell(row=row_num, column=7, value=item.get('unit', '斤'))
            ws.cell(row=row_num, column=8, value=item.get('unit_price', 0))
            ws.cell(row=row_num, column=9, value=item.get('subtotal', 0))
            ws.cell(row=row_num, column=10, value=item.get('remark', '') or (remark if item_idx == 0 else ''))
            row_num += 1

    # 调整列宽
    column_widths = [8, 15, 15, 25, 15, 10, 8, 10, 10, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 保存到内存流
    output = io.BytesIO()
    wb.save(output)
    return output


@reports_bp.route('/sales-summary', methods=['GET'])
def get_sales_summary():
    """
    获取销售汇总数据
    
    Query Params:
    - days: 统计天数(默认7天)
    """
    try:
        from database.db_config import get_db_session
        from models.models import Order, Customer
        from sqlalchemy import func
        
        days = request.args.get('days', 7, type=int)
        
        db = get_db_session()
        try:
            # 计算起始日期
            from datetime import timedelta
            end_date = date.today()
            start_date = end_date - timedelta(days=days-1)
            
            # 查询订单总数和总金额
            result = db.query(
                func.count(Order.id).label('total_orders'),
                func.sum(Order.total_amount).label('total_amount')
            ).filter(
                Order.order_date >= start_date,
                Order.order_date <= end_date
            ).first()
            
            total_orders = result.total_orders or 0
            total_amount = float(result.total_amount or 0)
            
            # 按日统计趋势
            daily_stats = db.query(
                Order.order_date,
                func.count(Order.id).label('order_count'),
                func.sum(Order.total_amount).label('daily_amount')
            ).filter(
                Order.order_date >= start_date,
                Order.order_date <= end_date
            ).group_by(Order.order_date).order_by(Order.order_date).all()
            
            trend_data = [{
                'date': stat.order_date.isoformat(),
                'order_count': stat.order_count,
                'amount': float(stat.daily_amount or 0)
            } for stat in daily_stats]
            
            return jsonify({
                'success': True,
                'summary': {
                    'total_orders': total_orders,
                    'total_amount': round(total_amount, 2),
                    'avg_daily_orders': round(total_orders / days, 1) if days > 0 else 0,
                    'avg_daily_amount': round(total_amount / days, 2) if days > 0 else 0
                },
                'trend': trend_data
            }), 200
            
        finally:
            db.close()
            
    except Exception as e:
        logger.error(f"获取销售汇总API异常: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/product-sales', methods=['GET'])
def get_product_sales():
    """
    获取商品销量排行
    
    Query Params:
    - limit: 返回数量(默认10)
    - days: 统计天数(默认30天)
    """
    try:
        from database.db_config import get_db_session
        from models.models import OrderItem, Product, Order
        from sqlalchemy import func
        from datetime import timedelta
        
        limit = request.args.get('limit', 10, type=int)
        days = request.args.get('days', 30, type=int)
        
        db = get_db_session()
        try:
            end_date = date.today()
            start_date = end_date - timedelta(days=days-1)
            
            # 查询商品销量排行
            results = db.query(
                Product.id,
                Product.product_name,
                func.sum(OrderItem.quantity).label('total_quantity'),
                func.count(OrderItem.id).label('order_count'),
                func.sum(OrderItem.subtotal).label('total_amount')
            ).join(
                OrderItem, Product.id == OrderItem.product_id
            ).join(
                Order, OrderItem.order_id == Order.id
            ).filter(
                Order.order_date >= start_date,
                Order.order_date <= end_date
            ).group_by(
                Product.id, Product.product_name
            ).order_by(
                func.sum(OrderItem.quantity).desc()
            ).limit(limit).all()
            
            ranking = [{
                'product_id': r.id,
                'product_name': r.product_name,
                'total_quantity': float(r.total_quantity or 0),
                'order_count': r.order_count,
                'total_amount': float(r.total_amount or 0)
            } for r in results]
            
            return jsonify({
                'success': True,
                'ranking': ranking
            }), 200
            
        finally:
            db.close()
            
    except Exception as e:
        logger.error(f"获取商品销量排行API异常: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/customer-activity', methods=['GET'])
def get_customer_activity():
    """
    获取客户活跃度统计
    
    Query Params:
    - days: 统计天数(默认30天)
    """
    try:
        from database.db_config import get_db_session
        from models.models import Order, Customer
        from sqlalchemy import func
        from datetime import timedelta
        
        days = request.args.get('days', 30, type=int)
        
        db = get_db_session()
        try:
            end_date = date.today()
            start_date = end_date - timedelta(days=days-1)
            
            # 查询客户活跃度
            results = db.query(
                Customer.id,
                Customer.customer_name,
                func.count(Order.id).label('order_count'),
                func.sum(Order.total_amount).label('total_amount'),
                func.max(Order.order_date).label('last_order_date')
            ).join(
                Order, Customer.id == Order.customer_id
            ).filter(
                Order.order_date >= start_date,
                Order.order_date <= end_date
            ).group_by(
                Customer.id, Customer.customer_name
            ).order_by(
                func.count(Order.id).desc()
            ).limit(20).all()
            
            activity_data = [{
                'customer_id': r.id,
                'customer_name': r.customer_name,
                'order_count': r.order_count,
                'total_amount': float(r.total_amount or 0),
                'last_order_date': r.last_order_date.isoformat() if r.last_order_date else None
            } for r in results]
            
            # 统计总体数据
            total_customers = db.query(func.count(Customer.id)).scalar()
            active_customers = len(activity_data)
            
            return jsonify({
                'success': True,
                'summary': {
                    'total_customers': total_customers,
                    'active_customers': active_customers,
                    'activity_rate': round((active_customers / total_customers * 100) if total_customers > 0 else 0, 2)
                },
                'customers': activity_data
            }), 200
            
        finally:
            db.close()
            
    except Exception as e:
        logger.error(f"获取客户活跃度API异常: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500
