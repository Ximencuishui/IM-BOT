"""
管理员专用API - 入口文件
提供用户管理、系统设置等管理员功能
已拆分为多个子模块，此文件作为统一入口
"""
import json
from flask import Blueprint, request, jsonify, g
from sqlalchemy.orm import Session
from database.db_config import get_db_session
from services.auth_service import admin_required, get_current_user_from_request
from models.user_models import User, License, PricingConfig, RuleTemplate, PaymentRecord, PricingPackage
from models.models import Order, OrderItem, Customer, Product
from datetime import datetime, date, timedelta
from utils.error_handler import handle_db_error, handle_error

admin_bp = Blueprint('admin', __name__, url_prefix='/api/admin')


def init_admin_routes():
    """初始化所有管理员路由"""
    from api.admin.users import init_users_routes
    from api.admin.permissions import init_permissions_routes
    from api.admin.audit import init_audit_routes, log_admin_action
    
    init_users_routes(admin_bp)
    init_permissions_routes(admin_bp)
    init_audit_routes(admin_bp)


@admin_bp.route('/licenses', methods=['GET'])
@admin_required
def get_all_licenses():
    """获取所有授权码（仅管理员）"""
    db: Session = get_db_session()
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        license_code = request.args.get('license_code')
        user_email = request.args.get('user_email')
        status = request.args.get('status')
        license_type = request.args.get('type')
        
        query = db.query(License)
        
        if license_code:
            query = query.filter(License.license_code.like(f'%{license_code}%'))
        if user_email:
            from models.user_models import User
            query = query.join(User).filter(User.email.like(f'%{user_email}%'))
        if status == 'active':
            query = query.filter(License.is_active == True, License.is_revoked == False)
        elif status == 'inactive':
            query = query.filter(License.is_active == False, License.is_revoked == False)
        elif status == 'revoked':
            query = query.filter(License.is_revoked == True)
        if license_type:
            query = query.filter(License.license_type == license_type)
        
        query = query.order_by(License.created_at.desc())
        
        total = query.count()
        licenses = query.offset((page - 1) * per_page).limit(per_page).all()
        
        return jsonify({
            'success': True,
            'licenses': [license.to_dict() for license in licenses],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_all_licenses')
    finally:
        db.close()


@admin_bp.route('/stats', methods=['GET'])
@admin_required
def get_system_stats():
    """获取系统统计信息（仅管理员）"""
    db: Session = get_db_session()
    try:
        total_users = db.query(User).count()
        active_users = db.query(User).filter(User.is_active == True).count()
        total_licenses = db.query(License).count()
        active_licenses = db.query(License).filter(License.is_active == True, License.is_revoked == False).count()
        revoked_licenses = db.query(License).filter(License.is_revoked == True).count()
        
        expiring_soon = db.query(License).filter(
            License.is_active == True,
            License.is_revoked == False,
            License.expires_at >= datetime.now(),
            License.expires_at <= datetime.now() + timedelta(days=7)
        ).count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total_users': total_users,
                'active_users': active_users,
                'total_licenses': total_licenses,
                'active_licenses': active_licenses,
                'revoked_licenses': revoked_licenses,
                'expiring_soon': expiring_soon,
                'timestamp': datetime.now().isoformat()
            }
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_system_stats')
    finally:
        db.close()


@admin_bp.route('/license-stats', methods=['GET'])
@admin_required
def get_license_stats():
    """获取授权码统计信息（仅管理员）"""
    db: Session = get_db_session()
    try:
        total_licenses = db.query(License).count()
        active_licenses = db.query(License).filter(
            License.is_active == True, 
            License.is_revoked == False
        ).count()
        revoked_licenses = db.query(License).filter(License.is_revoked == True).count()
        
        expiring_soon = db.query(License).filter(
            License.is_active == True,
            License.is_revoked == False,
            License.expires_at >= datetime.now(),
            License.expires_at <= datetime.now() + timedelta(days=7)
        ).count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total_licenses,
                'active': active_licenses,
                'revoked': revoked_licenses,
                'expiring': expiring_soon
            }
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_license_stats')
    finally:
        db.close()


@admin_bp.route('/orders', methods=['GET'])
@admin_required
def get_all_orders():
    """获取所有订单列表（仅管理员）"""
    db: Session = get_db_session()
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        customer_id = request.args.get('customer_id', type=int)
        status = request.args.get('status')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = db.query(Order)
        
        if customer_id:
            query = query.filter(Order.customer_id == customer_id)
        if status:
            query = query.filter(Order.status == status)
        if start_date:
            query = query.filter(Order.order_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
        if end_date:
            query = query.filter(Order.order_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
        
        query = query.order_by(Order.created_at.desc())
        
        total = query.count()
        orders = query.offset((page - 1) * per_page).limit(per_page).all()
        
        orders_data = []
        for order in orders:
            order_dict = order.to_dict()
            if order.customer:
                order_dict['customer_name'] = order.customer.customer_name
            orders_data.append(order_dict)
        
        return jsonify({
            'success': True,
            'orders': orders_data,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_all_orders')
    finally:
        db.close()


@admin_bp.route('/orders/<int:order_id>', methods=['GET'])
@admin_required
def get_order_detail(order_id):
    """获取订单详情（仅管理员）"""
    db: Session = get_db_session()
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'error': '订单不存在'}), 404
        
        order_dict = order.to_dict()
        if order.customer:
            order_dict['customer_name'] = order.customer.customer_name
        
        return jsonify({
            'success': True,
            'order': order_dict
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_order_detail')
    finally:
        db.close()


@admin_bp.route('/orders/<int:order_id>/status', methods=['PUT'])
@admin_required
def update_order_status(order_id):
    """更新订单状态（仅管理员）"""
    data = request.get_json()
    if not data or 'status' not in data:
        return jsonify({'success': False, 'error': '缺少状态参数'}), 400
    
    db: Session = get_db_session()
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            return jsonify({'success': False, 'error': '订单不存在'}), 404
        
        old_status = order.status
        order.status = data['status']
        if data.get('confirmed_by'):
            order.confirmed_by = data['confirmed_by']
        if data.get('remark'):
            order.remark = data['remark']
        
        db.commit()
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'order_update', f'更新订单状态: {order.id}, {old_status} -> {data["status"]}')
        
        return jsonify({
            'success': True,
            'message': f'订单状态从 {old_status} 更新为 {data["status"]}',
            'order': order.to_dict()
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'update_order_status')
    finally:
        db.close()


@admin_bp.route('/products', methods=['GET'])
@admin_required
def get_all_products():
    """获取所有商品列表（仅管理员）"""
    db: Session = get_db_session()
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        category = request.args.get('category')
        is_active = request.args.get('is_active')
        
        query = db.query(Product)
        
        if category:
            query = query.filter(Product.category == category)
        if is_active is not None:
            is_active_bool = is_active.lower() in ['true', '1', 'yes']
            query = query.filter(Product.is_active == (1 if is_active_bool else 0))
        
        query = query.order_by(Product.sort_order.asc(), Product.created_at.desc())
        
        total = query.count()
        products = query.offset((page - 1) * per_page).limit(per_page).all()
        
        return jsonify({
            'success': True,
            'products': [product.to_dict() for product in products],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_all_products')
    finally:
        db.close()


@admin_bp.route('/products', methods=['POST'])
@admin_required
def create_product():
    """创建新商品（仅管理员）"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400
    
    required_fields = ['product_name', 'unit', 'price']
    for field in required_fields:
        if field not in data:
            return jsonify({'success': False, 'error': f'缺少必要字段: {field}'}), 400
    
    db: Session = get_db_session()
    try:
        existing_product = db.query(Product).filter(Product.product_name == data['product_name']).first()
        if existing_product:
            return jsonify({'success': False, 'error': '该商品名称已存在'}), 400
        
        new_product = Product(
            product_name=data['product_name'],
            product_code=data.get('product_code', ''),
            shortcut_codes=','.join(data.get('shortcut_codes', [])) if isinstance(data.get('shortcut_codes'), list) else data.get('shortcut_codes', ''),
            unit=data['unit'],
            price=data['price'],
            category=data.get('category', ''),
            commission=data.get('commission', 0),
            points=data.get('points', 0),
            sort_order=data.get('sort_order', 0),
            is_active=data.get('is_active', 1)
        )
        
        attributes = data.get('attributes', [])
        if attributes:
            if len(attributes) > 10:
                return jsonify({'success': False, 'error': '商品属性不能超过10个'}), 400
            new_product.attributes = json.dumps(attributes, ensure_ascii=False)
        
        db.add(new_product)
        db.commit()
        db.refresh(new_product)
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'product_create', f'创建商品: {new_product.product_name}')
        
        return jsonify({
            'success': True,
            'product': new_product.to_dict()
        }), 201
    except Exception as e:
        return handle_db_error(e, db, 'create_product')
    finally:
        db.close()


@admin_bp.route('/products/<int:product_id>', methods=['PUT'])
@admin_required
def update_product(product_id):
    """更新商品信息（仅管理员）"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400
    
    db: Session = get_db_session()
    try:
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            return jsonify({'success': False, 'error': '商品不存在'}), 404
        
        if 'product_name' in data:
            existing = db.query(Product).filter(
                Product.product_name == data['product_name'],
                Product.id != product_id
            ).first()
            if existing:
                return jsonify({'success': False, 'error': '该商品名称已被其他商品使用'}), 400
            product.product_name = data['product_name']
        
        if 'product_code' in data:
            product.product_code = data['product_code']
        if 'shortcut_codes' in data:
            product.shortcut_codes = ','.join(data['shortcut_codes']) if isinstance(data['shortcut_codes'], list) else data['shortcut_codes']
        if 'unit' in data:
            product.unit = data['unit']
        if 'price' in data:
            product.price = data['price']
        if 'category' in data:
            product.category = data['category']
        if 'commission' in data:
            product.commission = data['commission']
        if 'points' in data:
            product.points = data['points']
        if 'sort_order' in data:
            product.sort_order = data['sort_order']
        if 'is_active' in data:
            product.is_active = data['is_active']
        
        if 'attributes' in data:
            attributes = data['attributes']
            if len(attributes) > 10:
                return jsonify({'success': False, 'error': '商品属性不能超过10个'}), 400
            product.attributes = json.dumps(attributes, ensure_ascii=False)
        
        db.commit()
        db.refresh(product)
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'product_update', f'更新商品: {product.product_name}')
        
        return jsonify({
            'success': True,
            'product': product.to_dict()
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'update_product')
    finally:
        db.close()


@admin_bp.route('/products/<int:product_id>', methods=['DELETE'])
@admin_required
def delete_product(product_id):
    """删除商品（仅管理员，软删除）"""
    db: Session = get_db_session()
    try:
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            return jsonify({'success': False, 'error': '商品不存在'}), 404
        
        product.is_active = 0
        db.commit()
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'product_delete', f'删除商品: {product.product_name}')
        
        return jsonify({
            'success': True,
            'message': '商品已删除'
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'delete_product')
    finally:
        db.close()


@admin_bp.route('/customers', methods=['GET'])
@admin_required
def get_all_customers():
    """获取所有客户列表（仅管理员）"""
    db: Session = get_db_session()
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        route_id = request.args.get('route_id', type=int)
        sales_person = request.args.get('sales_person')
        is_active = request.args.get('is_active')
        
        query = db.query(Customer)
        
        if route_id:
            query = query.filter(Customer.route_id == route_id)
        if sales_person:
            query = query.filter(Customer.sales_person == sales_person)
        if is_active is not None:
            is_active_bool = is_active.lower() in ['true', '1', 'yes']
            query = query.filter(Customer.is_active == (1 if is_active_bool else 0))
        
        query = query.order_by(Customer.created_at.desc())
        
        total = query.count()
        customers = query.offset((page - 1) * per_page).limit(per_page).all()
        
        customers_data = []
        for customer in customers:
            customer_dict = customer.to_dict()
            if customer.route:
                customer_dict['route_name'] = customer.route.route_name
            customers_data.append(customer_dict)
        
        return jsonify({
            'success': True,
            'customers': customers_data,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_all_customers')
    finally:
        db.close()


@admin_bp.route('/customers', methods=['POST'])
@admin_required
def create_customer():
    """创建新客户（仅管理员）"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400
    
    required_fields = ['customer_name']
    for field in required_fields:
        if field not in data:
            return jsonify({'success': False, 'error': f'缺少必要字段: {field}'}), 400
    
    db: Session = get_db_session()
    try:
        existing_customer = db.query(Customer).filter(Customer.customer_name == data['customer_name']).first()
        if existing_customer:
            return jsonify({'success': False, 'error': '该客户名称已存在'}), 400
        
        new_customer = Customer(
            customer_name=data['customer_name'],
            phone=data.get('phone', ''),
            address=data.get('address', ''),
            wx_group_id=data.get('wx_group_id', ''),
            wx_alias=data.get('wx_alias', ''),
            route_id=data.get('route_id'),
            sales_person=data.get('sales_person', ''),
            remark=data.get('remark', ''),
            is_active=data.get('is_active', 1)
        )
        
        db.add(new_customer)
        db.commit()
        db.refresh(new_customer)
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'customer_create', f'创建客户: {new_customer.customer_name}')
        
        return jsonify({
            'success': True,
            'customer': new_customer.to_dict()
        }), 201
    except Exception as e:
        return handle_db_error(e, db, 'create_customer')
    finally:
        db.close()


@admin_bp.route('/customers/<int:customer_id>', methods=['PUT'])
@admin_required
def update_customer(customer_id):
    """更新客户信息（仅管理员）"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400
    
    db: Session = get_db_session()
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            return jsonify({'success': False, 'error': '客户不存在'}), 404
        
        if 'customer_name' in data:
            existing = db.query(Customer).filter(
                Customer.customer_name == data['customer_name'],
                Customer.id != customer_id
            ).first()
            if existing:
                return jsonify({'success': False, 'error': '该客户名称已被其他客户使用'}), 400
            customer.customer_name = data['customer_name']
        
        if 'phone' in data:
            customer.phone = data['phone']
        if 'address' in data:
            customer.address = data['address']
        if 'wx_group_id' in data:
            customer.wx_group_id = data['wx_group_id']
        if 'wx_alias' in data:
            customer.wx_alias = data['wx_alias']
        if 'route_id' in data:
            customer.route_id = data['route_id']
        if 'sales_person' in data:
            customer.sales_person = data['sales_person']
        if 'remark' in data:
            customer.remark = data['remark']
        if 'is_active' in data:
            customer.is_active = data['is_active']
        
        db.commit()
        db.refresh(customer)
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'customer_update', f'更新客户: {customer.customer_name}')
        
        return jsonify({
            'success': True,
            'customer': customer.to_dict()
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'update_customer')
    finally:
        db.close()


@admin_bp.route('/customers/<int:customer_id>', methods=['DELETE'])
@admin_required
def delete_customer(customer_id):
    """删除客户（仅管理员，软删除）"""
    db: Session = get_db_session()
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            return jsonify({'success': False, 'error': '客户不存在'}), 404
        
        customer.is_active = 0
        db.commit()
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'customer_delete', f'删除客户: {customer.customer_name}')
        
        return jsonify({
            'success': True,
            'message': '客户已删除'
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'delete_customer')
    finally:
        db.close()


@admin_bp.route('/reports/sales-summary', methods=['GET'])
@admin_required
def get_sales_summary():
    """获取销售汇总报表（仅管理员）"""
    db: Session = get_db_session()
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            today = date.today()
            start_date = today.replace(day=1)
            end_date = today
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        orders = db.query(Order).filter(
            Order.order_date >= start_date,
            Order.order_date <= end_date,
            Order.status != 'cancelled'
        ).all()
        
        sales_stats = {}
        total_amount = 0
        total_orders = len(orders)
        
        for order in orders:
            customer = order.customer
            if not customer:
                continue
                
            sales_person = customer.sales_person or '未分配'
            
            if sales_person not in sales_stats:
                sales_stats[sales_person] = {
                    'sales_person': sales_person,
                    'order_count': 0,
                    'total_amount': 0.0,
                    'customer_count': set()
                }
            
            sales_stats[sales_person]['order_count'] += 1
            sales_stats[sales_person]['total_amount'] += float(order.total_amount)
            sales_stats[sales_person]['customer_count'].add(order.customer_id)
            total_amount += float(order.total_amount)
        
        for stats in sales_stats.values():
            stats['customer_count'] = len(stats['customer_count'])
        
        return jsonify({
            'success': True,
            'period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            },
            'summary': {
                'total_orders': total_orders,
                'total_amount': round(total_amount, 2),
                'sales_stats': list(sales_stats.values())
            }
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_sales_summary')
    finally:
        db.close()


@admin_bp.route('/reports/product-sales', methods=['GET'])
@admin_required
def get_product_sales_report():
    """获取商品销售报表（仅管理员）"""
    db: Session = get_db_session()
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = request.args.get('limit', 20, type=int)
        
        if not start_date or not end_date:
            today = date.today()
            start_date = today.replace(day=1)
            end_date = today
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        from sqlalchemy import func
        product_stats = db.query(
            OrderItem.product_name,
            func.sum(OrderItem.quantity).label('total_quantity'),
            func.sum(OrderItem.subtotal).label('total_amount'),
            func.count(OrderItem.id).label('order_times')
        ).join(Order).filter(
            Order.order_date >= start_date,
            Order.order_date <= end_date,
            Order.status != 'cancelled'
        ).group_by(
            OrderItem.product_name
        ).order_by(
            func.sum(OrderItem.quantity).desc()
        ).limit(limit).all()
        
        products_data = []
        for stat in product_stats:
            products_data.append({
                'product_name': stat.product_name,
                'total_quantity': float(stat.total_quantity),
                'total_amount': float(stat.total_amount),
                'order_times': stat.order_times
            })
        
        return jsonify({
            'success': True,
            'period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            },
            'products': products_data
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_product_sales_report')
    finally:
        db.close()


@admin_bp.route('/reports/customer-activity', methods=['GET'])
@admin_required
def get_customer_activity_report():
    """获取客户活跃度报表（仅管理员）"""
    db: Session = get_db_session()
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            today = date.today()
            start_date = today.replace(day=1)
            end_date = today
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        orders = db.query(Order).filter(
            Order.order_date >= start_date,
            Order.order_date <= end_date,
            Order.status != 'cancelled'
        ).all()
        
        customer_stats = {}
        total_orders = len(orders)
        
        for order in orders:
            customer_id = order.customer_id
            customer_name = order.customer.customer_name if order.customer else f'客户{customer_id}'
            
            if customer_id not in customer_stats:
                customer_stats[customer_id] = {
                    'customer_id': customer_id,
                    'customer_name': customer_name,
                    'order_count': 0,
                    'total_amount': 0.0
                }
            
            customer_stats[customer_id]['order_count'] += 1
            customer_stats[customer_id]['total_amount'] += float(order.total_amount)
        
        customers_data = sorted(customer_stats.values(), key=lambda x: x['order_count'], reverse=True)
        
        return jsonify({
            'success': True,
            'period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            },
            'summary': {
                'total_orders': total_orders,
                'active_customers': len(customers_data),
                'customers': customers_data
            }
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_customer_activity_report')
    finally:
        db.close()


@admin_bp.route('/reports/export/sales-summary', methods=['GET'])
@admin_required
def export_sales_summary():
    """导出销售汇总报表为Excel（仅管理员）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from flask import send_file
    import io
    
    db: Session = get_db_session()
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            today = date.today()
            start_date = today.replace(day=1)
            end_date = today
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        orders = db.query(Order).filter(
            Order.order_date >= start_date,
            Order.order_date <= end_date,
            Order.status != 'cancelled'
        ).all()
        
        sales_stats = {}
        total_amount = 0
        total_orders = len(orders)
        
        for order in orders:
            customer = order.customer
            if not customer:
                continue
                
            sales_person = customer.sales_person or '未分配'
            
            if sales_person not in sales_stats:
                sales_stats[sales_person] = {
                    'sales_person': sales_person,
                    'order_count': 0,
                    'total_amount': 0.0,
                    'customer_ids': set()
                }
            
            sales_stats[sales_person]['order_count'] += 1
            sales_stats[sales_person]['total_amount'] += float(order.total_amount)
            sales_stats[sales_person]['customer_ids'].add(order.customer_id)
            total_amount += float(order.total_amount)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "销售汇总"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells('A1:E1')
        title_cell = ws.cell(row=1, column=1, value=f"销售汇总报表 ({start_date.isoformat()} 至 {end_date.isoformat()})")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ['销售员', '订单数', '客户数', '销售额', '平均客单价']
        ws.append(headers)
        
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        row_num = 3
        for sales_person, stats in sorted(sales_stats.items()):
            customer_count = len(stats['customer_ids'])
            avg_order_amount = stats['total_amount'] / stats['order_count'] if stats['order_count'] > 0 else 0
            
            ws.cell(row=row_num, column=1, value=sales_person)
            ws.cell(row=row_num, column=2, value=stats['order_count'])
            ws.cell(row=row_num, column=3, value=customer_count)
            ws.cell(row=row_num, column=4, value=round(stats['total_amount'], 2))
            ws.cell(row=row_num, column=5, value=round(avg_order_amount, 2))
            row_num += 1
        
        ws.cell(row=row_num, column=1, value='总计')
        ws.cell(row=row_num, column=2, value=total_orders)
        ws.cell(row=row_num, column=4, value=round(total_amount, 2))
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=2).font = Font(bold=True)
        ws.cell(row=row_num, column=4).font = Font(bold=True)
        
        for i, width in enumerate([15, 10, 10, 15, 15], 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"销售汇总_{start_date.isoformat()}_至_{end_date.isoformat()}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return handle_db_error(e, db, 'export_sales_summary')
    finally:
        db.close()


@admin_bp.route('/reports/export/product-sales', methods=['GET'])
@admin_required
def export_product_sales():
    """导出商品销售报表为Excel（仅管理员）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from flask import send_file
    import io
    from sqlalchemy import func
    
    db: Session = get_db_session()
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = request.args.get('limit', 50, type=int)
        
        if not start_date or not end_date:
            today = date.today()
            start_date = today.replace(day=1)
            end_date = today
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        product_stats = db.query(
            OrderItem.product_name,
            func.sum(OrderItem.quantity).label('total_quantity'),
            func.sum(OrderItem.subtotal).label('total_amount'),
            func.count(OrderItem.id).label('order_times')
        ).join(Order).filter(
            Order.order_date >= start_date,
            Order.order_date <= end_date,
            Order.status != 'cancelled'
        ).group_by(
            OrderItem.product_name
        ).order_by(
            func.sum(OrderItem.quantity).desc()
        ).limit(limit).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "商品销售"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells('A1:E1')
        title_cell = ws.cell(row=1, column=1, value=f"商品销售报表 ({start_date.isoformat()} 至 {end_date.isoformat()})")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ['商品名称', '总数量', '总金额', '订单次数', '平均单价']
        ws.append(headers)
        
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        row_num = 3
        for stat in product_stats:
            avg_price = float(stat.total_amount) / float(stat.total_quantity) if float(stat.total_quantity) > 0 else 0
            
            ws.cell(row=row_num, column=1, value=stat.product_name)
            ws.cell(row=row_num, column=2, value=float(stat.total_quantity))
            ws.cell(row=row_num, column=3, value=float(stat.total_amount))
            ws.cell(row=row_num, column=4, value=stat.order_times)
            ws.cell(row=row_num, column=5, value=round(avg_price, 2))
            row_num += 1
        
        for i, width in enumerate([20, 12, 15, 12, 15], 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"商品销售_{start_date.isoformat()}_至_{end_date.isoformat()}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return handle_db_error(e, db, 'export_product_sales')
    finally:
        db.close()


@admin_bp.route('/pricing', methods=['GET'])
@admin_required
def get_pricing_config():
    """获取定价配置（仅管理员）"""
    db: Session = get_db_session()
    try:
        configs = db.query(PricingConfig).all()
        pricing = {c.config_key: c.to_dict() for c in configs}
        return jsonify({
            'success': True,
            'pricing': pricing
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_pricing_config')
    finally:
        db.close()


@admin_bp.route('/pricing', methods=['POST'])
@admin_required
def update_pricing_config():
    """更新定价配置（仅管理员）"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400
    
    db: Session = get_db_session()
    try:
        for key, value in data.items():
            config = db.query(PricingConfig).filter(PricingConfig.config_key == key).first()
            if config:
                config.config_value = value
            else:
                new_config = PricingConfig(
                    config_key=key,
                    config_value=value,
                    description=f'{key} 价格配置'
                )
                db.add(new_config)
        
        db.commit()
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'pricing_update', f'更新定价配置: {", ".join(data.keys())}')
        
        return jsonify({'success': True, 'message': '定价配置更新成功'}), 200
    except Exception as e:
        return handle_db_error(e, db, 'update_pricing_config')
    finally:
        db.close()


@admin_bp.route('/licenses/<int:license_id>/extend', methods=['POST'])
@admin_required
def admin_extend_license(license_id):
    """管理员手动展期授权码（支持其他渠道收款）"""
    data = request.get_json()
    if not data or 'months' not in data:
        return jsonify({'success': False, 'error': '请提供展期月数'}), 400
    
    db: Session = get_db_session()
    try:
        license_obj = db.query(License).filter(License.id == license_id).first()
        if not license_obj:
            return jsonify({'success': False, 'error': '授权码不存在'}), 404
        
        current_user = get_current_user_from_request(db)
        months = data['months']
        current_expiry = license_obj.expires_at or datetime.now()
        new_expiry = current_expiry + timedelta(days=30 * months)
        
        license_obj.expires_at = new_expiry
        license_obj.is_active = True
        
        payment = PaymentRecord(
            order_no=f"MANUAL_{datetime.now().strftime('%Y%m%d%H%M%S')}_{license_id}",
            user_id=license_obj.user_id,
            admin_id=current_user.id,
            amount=0,
            subject=f'管理员手动展期 {months} 个月',
            payment_method='manual',
            status='success',
            extend_info=json.dumps({'license_id': license_id, 'months': months}),
            paid_at=datetime.now()
        )
        db.add(payment)
        
        db.commit()
        db.refresh(license_obj)
        
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'license_extend', f'展期授权码: {license_obj.license_code}, {months}个月')
        
        return jsonify({
            'success': True,
            'message': f'授权码已展期 {months} 个月',
            'license': license_obj.to_dict()
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'admin_extend_license')
    finally:
        db.close()


@admin_bp.route('/licenses/<int:license_id>/revoke', methods=['POST'])
@admin_required
def admin_revoke_license(license_id):
    """管理员撤销授权码"""
    db: Session = get_db_session()
    try:
        license_obj = db.query(License).filter(License.id == license_id).first()
        if not license_obj:
            return jsonify({'success': False, 'error': '授权码不存在'}), 404
        
        if license_obj.is_revoked:
            return jsonify({'success': False, 'error': '授权码已被撤销'}), 400
        
        license_obj.is_active = False
        license_obj.is_revoked = True
        
        db.commit()
        db.refresh(license_obj)
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'license_revoke', f'撤销授权码: {license_obj.license_code}')
        
        return jsonify({
            'success': True,
            'message': '授权码已撤销',
            'license': license_obj.to_dict()
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'admin_revoke_license')
    finally:
        db.close()


@admin_bp.route('/licenses/batch-extend', methods=['POST'])
@admin_required
def admin_batch_extend_licenses():
    """管理员批量展期授权码"""
    data = request.get_json()
    if not data or 'license_ids' not in data or 'months' not in data:
        return jsonify({'success': False, 'error': '请提供授权码ID列表和展期月数'}), 400
    
    db: Session = get_db_session()
    try:
        license_ids = data['license_ids']
        months = data['months']
        
        licenses = db.query(License).filter(License.id.in_(license_ids)).all()
        
        if len(licenses) != len(license_ids):
            return jsonify({'success': False, 'error': '部分授权码不存在'}), 404
        
        current_user = get_current_user_from_request(db)
        success_count = 0
        failed_licenses = []
        
        for license_obj in licenses:
            try:
                current_expiry = license_obj.expires_at or datetime.now()
                new_expiry = current_expiry + timedelta(days=30 * months)
                
                license_obj.expires_at = new_expiry
                license_obj.is_active = True
                license_obj.is_revoked = False
                
                payment = PaymentRecord(
                    order_no=f"BATCH_{datetime.now().strftime('%Y%m%d%H%M%S')}_{license_obj.id}",
                    user_id=license_obj.user_id,
                    admin_id=current_user.id,
                    amount=0,
                    subject=f'管理员批量展期 {months} 个月',
                    payment_method='manual',
                    status='success',
                    extend_info=json.dumps({'license_id': license_obj.id, 'months': months}),
                    paid_at=datetime.now()
                )
                db.add(payment)
                success_count += 1
            except Exception as e:
                failed_licenses.append({
                    'license_id': license_obj.id,
                    'license_code': license_obj.license_code,
                    'error': str(e)
                })
        
        db.commit()
        
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'license_batch_extend', f'批量展期授权码: {success_count}个, {months}个月')
        
        message = f'成功展期 {success_count} 个授权码'
        if failed_licenses:
            message += f'，{len(failed_licenses)} 个失败'
        
        return jsonify({
            'success': True,
            'message': message,
            'success_count': success_count,
            'failed_count': len(failed_licenses),
            'failed_licenses': failed_licenses
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'admin_batch_extend_licenses')
    finally:
        db.close()


@admin_bp.route('/payments/alipay/create', methods=['POST'])
@admin_required
def create_alipay_order():
    """Admin 代付：创建支付宝订单"""
    from utils.alipay_utils import AliPayUtils
    data = request.get_json()
    if not data or not data.get('amount') or not data.get('subject'):
        return jsonify({'success': False, 'error': '缺少必要参数'}), 400

    db: Session = get_db_session()
    try:
        from config.settings import settings
        current_user = get_current_user_from_request(db)
        order_no = f"ADMIN_{datetime.now().strftime('%Y%m%d%H%M%S')}_{current_user.id}"
        
        payment = PaymentRecord(
            order_no=order_no,
            admin_id=current_user.id,
            user_id=data.get('user_id'),
            amount=data['amount'],
            subject=data['subject'],
            payment_method='alipay',
            status='pending',
            extend_info=json.dumps(data.get('extend_info', {}))
        )
        db.add(payment)
        db.commit()

        alipay = AliPayUtils()
        pay_url = alipay.create_web_pay_url(
            out_trade_no=order_no,
            total_amount=data['amount'],
            subject=data['subject'],
            return_url=f"{settings.FRONTEND_URL}/admin-portal?pay_result=success",
            notify_url=f"{settings.API_BASE_URL}/api/admin/payments/alipay/notify"
        )

        return jsonify({'success': True, 'pay_url': pay_url, 'order_no': order_no}), 200
    except Exception as e:
        return handle_db_error(e, db, 'create_alipay_order')
    finally:
        db.close()


@admin_bp.route('/payments/alipay/notify', methods=['POST'])
def alipay_notify():
    """支付宝异步通知回调"""
    from utils.alipay_utils import AliPayUtils
    data = request.form.to_dict()
    signature = data.pop('sign', '')
    
    db: Session = get_db_session()
    try:
        alipay = AliPayUtils()
        if alipay.verify_callback(data, signature):
            order_no = data.get('out_trade_no')
            trade_no = data.get('trade_no')
            
            payment = db.query(PaymentRecord).filter(PaymentRecord.order_no == order_no).first()
            if payment and payment.status != 'success':
                payment.status = 'success'
                payment.trade_no = trade_no
                payment.paid_at = datetime.now()
                
                db.commit()
                return 'success'
        return 'fail'
    except Exception as e:
        return handle_db_error(e, db, 'alipay_notify')
    finally:
        db.close()


@admin_bp.route('/rules/templates', methods=['GET'])
@admin_required
def get_all_rule_templates():
    """获取所有规则模板（仅管理员）"""
    db: Session = get_db_session()
    try:
        templates = db.query(RuleTemplate).all()
        return jsonify({
            'success': True,
            'templates': [t.to_dict() for t in templates]
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_all_rule_templates')
    finally:
        db.close()


@admin_bp.route('/rules/templates', methods=['POST'])
@admin_required
def create_rule_template():
    """创建规则模板（仅管理员）"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400
    
    required_fields = ['template_name', 'industry']
    for field in required_fields:
        if field not in data:
            return jsonify({'success': False, 'error': f'缺少必要字段: {field}'}), 400
    
    db: Session = get_db_session()
    try:
        new_template = RuleTemplate(
            template_name=data['template_name'],
            industry=data['industry'],
            description=data.get('description'),
            parse_rules=json.dumps(data.get('parse_rules', []), ensure_ascii=False),
            stat_rules=json.dumps(data.get('stat_rules', []), ensure_ascii=False),
            reply_rules=json.dumps(data.get('reply_rules', []), ensure_ascii=False),
            source_type=data.get('source_type', 'official'),
            is_public=data.get('is_public', True),
            is_featured=data.get('is_featured', False),
            is_active=data.get('is_active', True)
        )
        
        db.add(new_template)
        db.commit()
        db.refresh(new_template)
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'rule_template_create', f'创建规则模板: {new_template.template_name}')
        
        return jsonify({
            'success': True,
            'template': new_template.to_dict()
        }), 201
    except Exception as e:
        return handle_db_error(e, db, 'create_rule_template')
    finally:
        db.close()


@admin_bp.route('/rules/templates/<int:template_id>', methods=['PUT'])
@admin_required
def update_rule_template(template_id):
    """更新规则模板（仅管理员）"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400
    
    db: Session = get_db_session()
    try:
        template = db.query(RuleTemplate).filter(RuleTemplate.id == template_id).first()
        if not template:
            return jsonify({'success': False, 'error': '规则模板不存在'}), 404
        
        if 'template_name' in data:
            template.template_name = data['template_name']
        if 'industry' in data:
            template.industry = data['industry']
        if 'description' in data:
            template.description = data['description']
        if 'parse_rules' in data:
            template.parse_rules = json.dumps(data['parse_rules'], ensure_ascii=False)
        if 'stat_rules' in data:
            template.stat_rules = json.dumps(data['stat_rules'], ensure_ascii=False)
        if 'reply_rules' in data:
            template.reply_rules = json.dumps(data['reply_rules'], ensure_ascii=False)
        if 'source_type' in data:
            template.source_type = data['source_type']
        if 'is_public' in data:
            template.is_public = data['is_public']
        if 'is_featured' in data:
            template.is_featured = data['is_featured']
        if 'is_active' in data:
            template.is_active = data['is_active']
        
        db.commit()
        db.refresh(template)
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'rule_template_update', f'更新规则模板: {template.template_name}')
        
        return jsonify({
            'success': True,
            'template': template.to_dict()
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'update_rule_template')
    finally:
        db.close()


@admin_bp.route('/rules/templates/<int:template_id>', methods=['DELETE'])
@admin_required
def delete_rule_template(template_id):
    """删除规则模板（仅管理员，软删除）"""
    db: Session = get_db_session()
    try:
        template = db.query(RuleTemplate).filter(RuleTemplate.id == template_id).first()
        if not template:
            return jsonify({'success': False, 'error': '规则模板不存在'}), 404
        
        template.is_active = False
        db.commit()
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'rule_template_delete', f'删除规则模板: {template.template_name}')
        
        return jsonify({
            'success': True,
            'message': '规则模板已删除'
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'delete_rule_template')
    finally:
        db.close()


@admin_bp.route('/packages', methods=['GET'])
@admin_required
def get_all_packages():
    """获取所有定价套餐（仅管理员）"""
    db: Session = get_db_session()
    try:
        packages = db.query(PricingPackage).all()
        return jsonify({
            'success': True,
            'packages': [p.to_dict() for p in packages]
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'get_all_packages')
    finally:
        db.close()


@admin_bp.route('/packages', methods=['POST'])
@admin_required
def create_package():
    """创建定价套餐（仅管理员）"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400
    
    required_fields = ['package_name', 'price', 'duration_days']
    for field in required_fields:
        if field not in data:
            return jsonify({'success': False, 'error': f'缺少必要字段: {field}'}), 400
    
    db: Session = get_db_session()
    try:
        existing_package = db.query(PricingPackage).filter(PricingPackage.package_name == data['package_name']).first()
        if existing_package:
            return jsonify({'success': False, 'error': '该套餐名称已存在'}), 400
        
        new_package = PricingPackage(
            package_name=data['package_name'],
            description=data.get('description', ''),
            price=data['price'],
            duration_days=data['duration_days'],
            features=json.dumps(data.get('features', []), ensure_ascii=False),
            is_active=data.get('is_active', True),
            sort_order=data.get('sort_order', 0)
        )
        
        db.add(new_package)
        db.commit()
        db.refresh(new_package)
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'package_create', f'创建套餐: {new_package.package_name}')
        
        return jsonify({
            'success': True,
            'package': new_package.to_dict()
        }), 201
    except Exception as e:
        return handle_db_error(e, db, 'create_package')
    finally:
        db.close()


@admin_bp.route('/packages/<int:package_id>', methods=['PUT'])
@admin_required
def update_package(package_id):
    """更新定价套餐（仅管理员）"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': '请求数据为空'}), 400
    
    db: Session = get_db_session()
    try:
        package = db.query(PricingPackage).filter(PricingPackage.id == package_id).first()
        if not package:
            return jsonify({'success': False, 'error': '套餐不存在'}), 404
        
        if 'package_name' in data:
            existing = db.query(PricingPackage).filter(
                PricingPackage.package_name == data['package_name'],
                PricingPackage.id != package_id
            ).first()
            if existing:
                return jsonify({'success': False, 'error': '该套餐名称已被其他套餐使用'}), 400
            package.package_name = data['package_name']
        
        if 'description' in data:
            package.description = data['description']
        if 'price' in data:
            package.price = data['price']
        if 'duration_days' in data:
            package.duration_days = data['duration_days']
        if 'features' in data:
            package.features = json.dumps(data['features'], ensure_ascii=False)
        if 'is_active' in data:
            package.is_active = data['is_active']
        if 'sort_order' in data:
            package.sort_order = data['sort_order']
        
        db.commit()
        db.refresh(package)
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'package_update', f'更新套餐: {package.package_name}')
        
        return jsonify({
            'success': True,
            'package': package.to_dict()
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'update_package')
    finally:
        db.close()


@admin_bp.route('/packages/<int:package_id>', methods=['DELETE'])
@admin_required
def delete_package(package_id):
    """删除定价套餐（仅管理员，软删除）"""
    db: Session = get_db_session()
    try:
        package = db.query(PricingPackage).filter(PricingPackage.id == package_id).first()
        if not package:
            return jsonify({'success': False, 'error': '套餐不存在'}), 404
        
        package.is_active = False
        db.commit()
        
        current_user = get_current_user_from_request(db)
        if current_user:
            from api.admin.audit import log_admin_action
            log_admin_action(db, current_user.id, current_user.username,
                           'package_delete', f'删除套餐: {package.package_name}')
        
        return jsonify({
            'success': True,
            'message': '套餐已删除'
        }), 200
    except Exception as e:
        return handle_db_error(e, db, 'delete_package')
    finally:
        db.close()


init_admin_routes()
